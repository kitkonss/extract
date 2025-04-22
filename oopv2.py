# app.py  –  Transformer OCR ▸ POWTR‑CODE ▸ MxLoader (.xlsm)
import streamlit as st, pandas as pd, requests, json, base64, io, re, os
from openpyxl import load_workbook
from pathlib import Path

# ───────────── 0. LOAD FIXED FILES ──────────────────────────────────────────
TPL   = Path('Template-MxLoader-Classification POW-TR.xlsm')
ATTR  = Path('ATTRIBUTE.xlsx')

if not TPL.exists():
    TPL = st.file_uploader('📂 template .xlsm', ['xlsm'])
    if TPL is None: st.stop()
wb = load_workbook(TPL, keep_vba=True)

if not ATTR.exists():
    ATTR = st.file_uploader('📑 ATTRIBUTE.xlsx', ['xlsx','xls'])
    if ATTR is None: st.stop()

# ───────────── 1. UI ───────────────────────────────────────────────────────
st.title('⚡ Transformer OCR → POWTR‑CODE → MxLoader (.xlsm)')

pam_xls = st.file_uploader('📒 PAM.xlsx', ['xlsx','xls'])
imgs    = st.file_uploader('🖼️ Nameplate images', ['jpg','jpeg','png'],
                           accept_multiple_files=True)
api_key = os.getenv('GEMINI_API_KEY') or st.text_input('🔑 Gemini API key', type='password')

if pam_xls is not None:
    pam_df  = pd.read_excel(pam_xls)
    loc_col = st.selectbox('คอลัมน์ Location/AssetNUM ใน PAM', pam_df.columns,
                           index=list(pam_df.columns).index('Location') if 'Location' in pam_df.columns else 0)
else:
    pam_df  = pd.DataFrame(); loc_col=''

loc_map = {}
if imgs and not pam_df.empty:
    st.markdown('**กรอก Location/AssetNUM ให้รูปแต่ละไฟล์**')
    for im in imgs:
        loc_map[im.name] = st.text_input(im.name, key=im.name)

site_default = st.text_input('SITEID (default)', 'SBK0')

# ───────────── 2. PROMPT & ATTRIBUTE MAP ───────────────────────────────────
def build_prompt(attr_path):
    df = pd.read_excel(attr_path, header=None)
    df.columns = ['attr'] + df.columns[1:].tolist()
    attrs = [str(a).strip() for a in df['attr'] if str(a).strip()]
    idx_map = {str(i+1): a for i, a in enumerate(attrs)}

    prompt = """
คืนข้อมูลเป็น JSON ตามโครง:
{
  "HIGH_SIDE_VOLTAGE_KV": <kV>,
  "PHASE": <1 หรือ 3>,
  "COOLING_TYPE": "<ONAN/ONAF/OFAF/DRY...>",
  "TAP_CHANGER": "ON‑LOAD" | "OFF‑CIRCUIT" | "—",
  "VECTOR_GROUP": "<YnD11 ฯลฯ>"
}
ถ้าไม่พบค่า ใส่ค่าว่าง \"\" (ห้ามเดา) และอย่าใส่ค่าจาก BIL หรือ AC withstand
นอกจากนี้ให้ดึงค่าต่อไปนี้ (หากหาเจอ) ในรูปแบบ key เป็นหมายเลข index:
"""
    for i, a in enumerate(attrs, 1):
        prompt += f"{i}: {a}\n"
    prompt += "หากไม่พบให้ใส่ \"-\""

    return prompt, idx_map

prompt, idx_map = build_prompt(ATTR)

# ───────────── 3. OCR (Gemini) ─────────────────────────────────────────────
def encode_img(file): return base64.b64encode(file.getvalue()).decode('utf-8')

def gemini_ocr(key, img_b64, prompt_text):
    url = f"https://generativelanguage.googleapis.com/v1/models/gemini-2.0-flash:generateContent?key={key}"
    body = {
        "contents":[{"parts":[{"text":prompt_text},
                              {"inline_data":{"mime_type":"image/jpeg","data":img_b64}}]}],
        "generation_config":{"temperature":0.2,"max_output_tokens":4096}
    }
    r = requests.post(url, json=body)
    if r.status_code != 200:
        return {"error": f"{r.status_code}: {r.text}"}
    txt = r.json()['candidates'][0]['content']['parts'][0]['text']
    try:
        return json.loads(txt[txt.find('{'): txt.rfind('}')+1])
    except Exception:
        return {"raw_text": txt}

# ───────────── 4. POWTR‑CODE LOGIC ─────────────────────────────────────────
oil_kw = {'OIL','ONAN','ONAF','OFAF','OFWF','OA','OF','ON','ONO','OFA'}

def kV_detect(dic):
    """
    • รวมตัวเลขที่คั่น space/comma ('241 500') → 241500
    • หน่วย V → /1000, ไม่มีหน่วยแต่ >1000 → assume Volt
    • ตัด BIL / AC / IMPULSE และ >765kV
    """
    good=[]; pat=re.compile(r'(\d{2,7}(?:[ ,]\d{3})*(?:[.,]\d+)?)\s*(kV|KV|kv|V|v)?')
    for txt in map(str, dic.values()):
        up=txt.upper()
        if any(x in up for x in ('BIL','/ AC',' AC ','IMPULSE','LIGHTNING')): continue
        for raw,unit in pat.findall(txt):
            num=raw.replace(' ','').replace(',','')
            try: val=float(num.replace(',','.'))
            except: continue
            kv = val/1000 if (unit and unit.lower().startswith('v') or (not unit and val>1000)) else val
            if kv<=765: good.append(kv)
    return (max(good), good) if good else (None, [])

def gen_powtr(data):
    phase = str(data.get('PHASE','3')).replace('.0','')
    kv, cand = kV_detect(data)
    v_char = '-' if kv is None else ('E' if kv>=345 else 'H' if kv>=100 else 'M' if kv>=1 else 'L')
    cooling_field = (str(data.get('COOLING_TYPE','')) + ' ' + str(data.get('TYPE OF COOLING',''))).upper()
    t_char = 'O' if any(k in cooling_field for k in oil_kw) else 'D'
    tap_char = 'O' if 'ON' in str(data.get('TAP_CHANGER','')).upper() else 'F'
    return f'POWTR-{phase}{v_char}{t_char}{tap_char}', kv, cand

# ───────────── 5. SHEET HELPER ────────────────────────────────────────────
ASHEET,HROW,DSTART = 'AssetAttr', 2, 3
header = [c.value for c in wb[ASHEET][HROW]]
col = {h:i for i,h in enumerate(header) if h}

def idx2attr(k): return idx_map.get(str(k).strip(), k)

def blank(val, attr):
    """ค่าว่างเมื่อ OCR คืน '-', '', None หรือสะกดซ้ำ attribute"""
    if val in {'-','',None}: return ''
    if str(val).strip().upper() == str(attr).strip().upper():
        return ''
    return val

def build_rows(assetnum, siteid, powtr_code, ocr_dict):
    rows=[]
    hier=f"POWTR \\ {powtr_code}"
    for k,v in ocr_dict.items():
        if k in ('error','raw_text'): continue
        attr = idx2attr(k)
        value = blank(v, attr)
        m = re.search(r'\((.*?)\)\s*$', attr)
        unit = m.group(1).strip() if m else ''
        r = ['']*len(header)
        r[col['ASSETNUM']], r[col['SITEID']], r[col['HIERARCHYPATH']] = assetnum, siteid, hier
        r[col['ASSETSPEC.\nASSETATTRID']]  = attr
        r[col['ASSETSPEC.ALNVALUE']]       = value
        r[col['ASSETSPEC.MEASUREUNITID']]  = unit
        rows.append(r)
    return rows

def show_debug(idx, ocr_dict, kv, cand):
    with st.expander(f'Debug – image #{idx+1}'):
        st.json(ocr_dict)
        st.write('kV candidates ⇒', cand)
        st.write('chosen kV ⇒', kv if kv else 'N/A')

# ───────────── 6. RUN ─────────────────────────────────────────────────────
if st.button('🚀 Run') and api_key and imgs and not pam_df.empty:
    ws = wb[ASHEET]
    if ws.max_row >= DSTART:
        ws.delete_rows(DSTART, ws.max_row-DSTART+1)

    results=[]; prog=st.progress(0.)
    for i, im in enumerate(imgs, 1):
        prog.progress(i/len(imgs))
        loc = loc_map.get(im.name,'').strip()
        if not loc:
            st.warning(f'{im.name} – ไม่มี Location'); continue

        ocr_dict = gemini_ocr(api_key, encode_img(im), prompt)
        powtr_code, kv_used, kv_list = gen_powtr(ocr_dict if isinstance(ocr_dict,dict) else {})
        show_debug(i, ocr_dict, kv_used, kv_list)

        pam_cls = pam_df.loc[pam_df[loc_col]==loc, 'Classification'].iat[0] \
                  if loc in pam_df[loc_col].values and 'Classification' in pam_df else ''
        results.append({'Image':im.name,'Asset':loc,
                        'POWTR(OCR)':powtr_code,'Classification(PAM)':pam_cls,
                        'Match?':powtr_code == pam_cls})

        for r in build_rows(loc, site_default, powtr_code, ocr_dict):
            ws.append(r)

    st.subheader('ผลการตรวจ')
    st.dataframe(pd.DataFrame(results))

    buff = io.BytesIO(); wb.save(buff); buff.seek(0)
    st.download_button('⬇️ Download MxLoader file', buff,
                       'MxLoader_POWTR_Result.xlsm',
                       'application/vnd.ms-excel.sheet.macroEnabled.12')
